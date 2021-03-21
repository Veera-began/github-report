import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import messagebox
import openpyxl
import tkinter
from tkinter import *
import os
import win32com.client as win32
path_of_file = os.path.dirname(__file__)
new_path = os.path.dirname(os.path.abspath(__file__))
print(path_of_file)
current_path = os.getcwd()

window = tk.Tk()
window.geometry('800x400')
window.title("Automation Sprint Plan Generator")
#########################################################################################################
frame_1 = ttk.Frame()
lb1 = ttk.Label(master=frame_1,
                text="Follow Below Steps to Generate Automaion Sprint Plan", foreground="white", background="black")
# , foreground="white", background="black"
lb1.pack()
# l.grid(column=0, row=0)
lb2 = ttk.Label(master=frame_1, relief=RAISED,
                text="S1.Will extract all test cases and Planned Test cases for each platform from Sprint Plan workook to a new Temp workook.", background="#2F4F4F", foreground="white")
# lb2.config(anchor=LEFT)
lb2.pack()
#lb2.place(x=0, y=20)
# lb2.place(anchor=E)
# frame_1.pack()
# l1.grid(column=0, row=1)
# , background="#2F4F4F"

#######################################################################################################
frame_2 = ttk.Frame()
lb3 = ttk.Label(master=frame_2, text="Enter the Name of Sprint Plan Worbook")
lb3.grid(column=0, row=0)
# lbl.pack()
plan_workook = ttk.Entry(master=frame_2, width=40)
plan_workook.grid(column=1, row=0)
# plan_workook.pack()
lb4 = ttk.Label(
    master=frame_2, text="Enter Some Random Temp workbook name eg:Veera.xlsx format")
lb4.grid(column=0, row=1)
# 2.pack()
temp = ttk.Entry(master=frame_2, width=40)
temp.grid(column=1, row=1)
# temp.pack()

########################################################
#Will exract all tcs into temp sheet
#####################################################
def testcases():
    wb1 = openpyxl.Workbook()
    wb1.save(filename=current_path + "\\" + temp.get())
    wb = load_workbook(
        filename=current_path + "\\" + plan_workook.get())
    plan = wb["Automation"]
    new_sheet = wb1.create_sheet("Testcases with Compoent")
    new_sheet["A1"] = "Test Script ID"
    new_sheet["B1"] = "Component"
    cur_automation_row = 2
    new_sheet_row = 2
    automation_rows = plan.max_row
    while cur_automation_row <= automation_rows:
        component = plan.cell(row=cur_automation_row, column=7).value
        test_script_id = plan.cell(row=cur_automation_row, column=6).value
        if type(test_script_id) == str:
            if test_script_id.find(",") >= 0:
                mul_tcs = []
                mul_tcs = test_script_id.split(",")
                for each in mul_tcs:
                    new_sheet.cell(row=new_sheet_row, column=1).value = each
                    new_sheet.cell(row=new_sheet_row,
                                   column=2).value = component
                    new_sheet_row += 1
            else:
                new_sheet.cell(row=new_sheet_row,
                               column=1).value = test_script_id
                new_sheet.cell(row=new_sheet_row, column=2).value = component
                new_sheet_row += 1
        cur_automation_row += 1
    wb1.save(filename=current_path + "\\" +
             temp.get())

###########################################################################
#Will extract planned TCs across all platforms
##########################################################################
def extract():
    # print(txt.get(), txt1.get())
    testcases()
    #wb1 = openpyxl.Workbook()
    wb1 = load_workbook(filename=current_path + "\\" + temp.get())
    wb = load_workbook(
        filename=current_path + "\\" + plan_workook.get())
    # s1=wb.active
    # print(s1)
    print(wb.sheetnames)
    plan = wb["Automation"]
    for row in plan.iter_rows(max_row=1, values_only=True):
        fr = row
    fr = list(fr)
    start = fr.index("AXB6")
    end = start+8
    for platform in range(start, end):
        print(fr[platform])
        for col in plan.iter_cols(min_col=platform+1, max_col=platform+1, values_only=True):
            planned = col
        # print(planned)
        planned = planned[1:]
        new_sheet = wb1.create_sheet(fr[platform])
        new_sheet["A1"] = "Test Script ID"
        new_row = 2
        old_row = 1
        for i in planned:
            if i == "Yes" or i == "Planned":
                # print(plan.cell(row=old_row+1,column=6).value)
                if type(plan.cell(row=old_row+1, column=6).value) == str:
                    new_sheet.cell(row=new_row, column=1).value = plan.cell(
                        row=old_row+1, column=6).value
                    new_row += 1
            old_row += 1
    wb1.save(filename=current_path + "\\" +
             temp.get())

    messagebox.showinfo('Message title', 'Succesfully done S1')

#####################################################################################################


frame_3 = ttk.Frame()
btn1 = ttk.Button(master=frame_3, text="Proceed1",
                  command=extract)
btn1.pack()


lb5 = ttk.Label(master=frame_3, relief=RAISED, justify=LEFT,
                text="S2.Will make sure that each cell has only one Test Case and Removes Dulpicates platform wise in temp workook.", foreground="white", background="#2F4F4F")
lb5.pack()


def Remove_Dup():
    wb = load_workbook(filename=current_path + "\\" +
                       temp.get())
    sheets = wb.sheetnames
    sheets = sheets[2:]
    for each_sheet in sheets:
        current_sheet = wb[each_sheet]
        for col in current_sheet.iter_cols(max_col=1, values_only=True):
            tcs = col
        tcs = tcs[1:]
        # print(tcs)
        r = 2
        for each in tcs:
            var = each.find(",")
            print(each)
            print(var)
            multi_tcs = []
            if var > 0:
                multi_tcs = each.split(",")
                print(multi_tcs)
                current_sheet.cell(row=r, column=1).value = multi_tcs[0]
                print(current_sheet.cell(row=r, column=1).value)
                r += 1
                print(r, "  inside")
                current_sheet.insert_rows(idx=r, amount=len(multi_tcs)-1)
                for i in range(1, len(multi_tcs)):
                    print(i)
                    current_sheet.cell(row=r, column=1).value = multi_tcs[i]
                    print(current_sheet.cell(row=r, column=1).value)
                    print(multi_tcs[i])
                    r += 1
                    print(r, " ineer loop")
            else:
                r += 1
                print(r, " outside")
        for col in current_sheet.iter_cols(max_col=1, values_only=True):
            tcs = col
        tcs = tcs[1:]
        current_sheet.delete_cols(idx=1)
    # print(cols)
        unique_tc = []
        for each in tcs:
            if each in unique_tc:
                pass
            else:
                unique_tc.append(each)
        r = 2
        current_sheet.cell(row=1, column=1).value = each_sheet + \
            "(" + str(len(unique_tc))+")"
        for each in unique_tc:
            current_sheet.cell(row=r, column=1).value = each
            r += 1
    wb.save(filename=current_path + "\\" +
            temp.get())
    messagebox.showinfo('Message title', 'Succesfully done S2')


btn2 = ttk.Button(master=frame_3, text="Proceed2",
                  command=Remove_Dup)
btn2.pack()

lb6 = ttk.Label(master=frame_3, relief=RAISED, justify=LEFT,
                text="S3.Will map component to all unique TCs to new \"Unique_TC\" sheet in Planned workbook from temp.xlsx workbook.", foreground="white", background="#2F4F4F")
lb6.pack()


def Copy_Unique():
    wb_from = load_workbook(filename=current_path + "\\" +
                            temp.get())
    wb_to = load_workbook(
        filename=current_path + "\\" + plan_workook.get())
    sheets = wb_from.sheetnames
    sheets = sheets[1:]
    print(sheets)
    # new_sheet=wb1.create_sheet(fr[platform])
    Unique = wb_to.create_sheet("Unique_TC")
    planned_tcs = []
    component_sheet = wb_from[sheets[0]]
 
    all_unique_tc = []
    compo_row=2
    compo_max=component_sheet.max_row
    while compo_row<=compo_max:
        t=component_sheet.cell(row=compo_row,column=1).value
        if t in all_unique_tc:
            pass
        else:
            all_unique_tc.append(t)
        compo_row+=1
    tc_with_component_rows = component_sheet.max_row
    Unique.cell(row=1, column=1).value = "Test Script ID"
    Unique.cell(row=1, column=2).value = "Component"
    #Compoent mapping to testcases
    r = 2
    for each in all_unique_tc:
        Unique.cell(row=r, column=1).value = each
        final_comp = []
        cur_row = 2
        while cur_row <= tc_with_component_rows:
            tc = component_sheet.cell(row=cur_row, column=1).value
            comp = component_sheet.cell(row=cur_row, column=2).value
            if tc == each:
                #final_comp = final_comp+"," + comp
                final_comp.append(comp)
            cur_row += 1
        final_comp = set(final_comp)
        final_comp = list(final_comp)
        final_comp = ",".join(final_comp)
        Unique.cell(row=r, column=2).value = final_comp
        r += 1
    wb_to.save(
        filename=current_path + "\\" + plan_workook.get())
    #mapping planed testcases across all platforms
    #print(all_unique_tc)
    sheets = sheets[1:]
    unique_rows = Unique.max_row
    c = 4
    for each_sheet in sheets:
        print(each_sheet)
        current_sheet = wb_from[each_sheet]
        print(current_sheet)
        tcs = []
        for col in current_sheet.iter_cols(max_col=1, min_col=1, values_only=True):
            tcs = col

        Unique.cell(row=1, column=c).value = tcs[0]
        tcs = list(tcs[1:])
        print(tcs)
        for each_tc in tcs:
            indx = all_unique_tc.index(each_tc)
            Unique.cell(row=indx+2, column=c).value = "Yes"
        c += 1
    wb_to.save(
        filename=current_path + "\\" + plan_workook.get())
    messagebox.showinfo('Message title', 'Succesfully done S3')


btn3 = ttk.Button(master=frame_3, text="Proceed3",
                  command=Copy_Unique)
btn3.pack()

lb7 = ttk.Label(master=frame_3, relief=RAISED, justify=LEFT,
                text="S4.Will convert xls workook to xlsx format and will map each TCs to respective Test Type and Applicable or Not Applicabe for respective platform.", foreground="white", background="#2F4F4F")
lb7.pack()
###########################################################################################################
frame_4 = ttk.Frame()
lb8 = ttk.Label(
    master=frame_4, text="Enter the Name of Script Mapping Workook i.e downloaded from Automaics")
lb8.grid(column=0, row=0)
# lbl.pack()
script_mapping_from = ttk.Entry(master=frame_4, width=40)
script_mapping_from.grid(column=1, row=0)
lb9 = ttk.Label(master=frame_4,
                text="Enter the Conversion workook name with xlsx format eg:Veera.xlsx", foreground="black")
lb9.grid(column=0, row=1)
script_mapping_to = ttk.Entry(master=frame_4, width=40)
script_mapping_to.grid(column=1, row=1)


def Conversion():
    file_name = current_path + "\\" + script_mapping_from.get()
    print(file_name)
    final_name = current_path + "\\" + script_mapping_to.get()
    print(final_name)
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(file_name)
    wb.SaveAs(final_name, FileFormat=51)
    wb.Close()
    messagebox.showinfo(
        'Message title', 'Succesfully Converted to xlsx format')


def Mapping():
    Conversion()
    wb_from = load_workbook(
        filename=current_path + "\\" + plan_workook.get())
    wb_to = load_workbook(filename=current_path + "\\" +
                          script_mapping_to.get())
    print(wb_to.sheetnames)
    #mapping Testtype of component 
    map_sheet = wb_to["Sheet1"]
    Auto_tcs=[]
    for col in map_sheet.iter_cols(max_col=1, min_col=1, values_only=True):
        col = col
    Auto_tcs = list(col)
    planned_sheet = wb_from["Unique_TC"]
    planned_sheet.cell(row=1,column=3).value="Test Type"
    unq_row=2
    unq_max=planned_sheet.max_row
    t=""
    while unq_row<=unq_max:
        t=planned_sheet.cell(row=unq_row,column=1).value
        if t in Auto_tcs:
            indx=Auto_tcs.index(t)
            planned_sheet.cell(row=unq_row,column=3).value=map_sheet.cell(row=indx,column=5).value
        else:
            planned_sheet.cell(row=unq_row,column=3).value="NA"
        unq_row+=1
    wb_from.save(
        filename=current_path + "\\" + plan_workook.get())
    #mapping A/NA for planned TCs accross all platforms
    map_App = wb_to["Script_Mapping"]
    for col in map_App.iter_cols(max_col=1, min_col=1, values_only=True):
        all_TC = col
    for row in map_App.iter_rows(max_row=1, min_row=1, values_only=True):
        headings = row
    run_on_models_col=headings.index("RUN ON MODELS")+1
    print(run_on_models_col)
    platforms = ["ARRIS-XB6", "TECH-XB6", "CISCO-XB3",
                 "ARRIS-XB3", "PACE-XF3", "PACE-CFG3", "TECH-CBR", "TECH-XB7"]
    # print(Auto_tcs)
    
    # print(wb_from.sheetnames)
    overall_column = 4
    for platform in platforms:
        print("*************Starting"+platform +
              "****************************************")
        for each in planned_sheet.iter_cols(max_col=overall_column, min_col=overall_column, values_only=True):
            plan_tc = each
        plan_tc=list(plan_tc)
        print(plan_tc)
        li=[]
        li=plan_tc[0].split("(")
        plan_tc=plan_tc[1:]
        planned_sheet.insert_cols(idx=overall_column+1, amount=1)
        r = 2
        counter = 0
        for each in plan_tc:
            tc=""
            if type(each)==str and each=="Yes":
                tc=planned_sheet.cell(row=r,column=1).value
                test_type=planned_sheet.cell(row=r,column=3).value
                if tc in all_TC and test_type != "NA":
                    ind = all_TC.index(tc)
                    final = ind+1
                    print(final)
                    print(map_App.cell(row=final, column=1).value)
                    Run_models = map_App.cell(row=final, column=run_on_models_col).value
                    # print(Run_models)
                    # print(Run_models)
                    print(type(Run_models))
                    final_list_platform = []
                    if type(Run_models) == str:
                        final_list_platform = list(Run_models.split("\n"))
                        print(final_list_platform)
                        if platform in final_list_platform:
                            planned_sheet.cell(row=r, column=overall_column+1).value = "A"
                            counter += 1
                        else:
                            planned_sheet.cell(row=r, column=overall_column+1).value = "NA"
                else:
                    planned_sheet.cell(row=r, column=overall_column+1).value = "NA"
                if test_type == "NA":
                    planned_sheet.cell(row=r, column=overall_column+1).value = "NA"
            r += 1
        planned_sheet.cell(row=1, column=overall_column +
                           1).value = "A in "+li[0] + "(" + str(counter) + ")"
        overall_column = overall_column+2
    wb_from.save(
        filename=current_path + "\\" + plan_workook.get())
    messagebox.showinfo(
        'Message title', 'Succesfully done S4 and Generated Automation Plan Successfully')
    window.destroy()


frame_5 = ttk.Frame()
btn4 = ttk.Button(master=frame_5, text="Proceed4",
                  command=Mapping)
btn4.pack()
####################################################################################################################
frame_1.pack()
frame_2.pack()
frame_3.pack()
frame_4.pack()
frame_5.pack()
window.mainloop()
